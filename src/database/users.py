import logging

import openpyxl
from openpyxl.styles import Font, PatternFill

from constants import (USERS_FILENAME, USERS_FILL_TYPE, USERS_HEADER_COLOR,
                       USERS_HEADER_FONT_COLOR, USERS_HEADERS, USERS_TITLE)
from src.database.base import db_manager
from src.database.merch import is_got_any_merch


async def create_users_table():
    """
    Create the `users` table in the database if it does not already exist.

    The table schema is:
        - id: integer primary key
        - username: text
        - quest_started: integer flag (0 or 1)
        - quest1_points … quest11_points: integer counters for each quest step
        - quest_station: integer current station number
        - quest_points: integer total quest points
        - quize_points: integer total quiz points
        - quize_1 … quize_5: integer counters for each quiz question

    This function will silently do nothing if the table already exists,
    and will print an error message if any exception occurs.
    """
    try:
        async with db_manager.get_connection() as conn:
            await conn.execute("""
                CREATE TABLE IF NOT EXISTS users (
                    id INTEGER PRIMARY KEY,
                    username TEXT,
                    quest_started INTEGER DEFAULT 0,
                    quest1_points INTEGER DEFAULT 0,
                    quest2_points INTEGER DEFAULT 0,
                    quest3_points INTEGER DEFAULT 0,
                    quest4_points INTEGER DEFAULT 0,
                    quest5_points INTEGER DEFAULT 0,
                    quest6_points INTEGER DEFAULT 0,
                    quest7_points INTEGER DEFAULT 0,
                    quest8_points INTEGER DEFAULT 0,
                    quest9_points INTEGER DEFAULT 0,
                    quest10_points INTEGER DEFAULT 0,
                    quest11_points INTEGER DEFAULT 0,
                    quest_station INTEGER DEFAULT 0,
                    quest_points INTEGER DEFAULT 0,
                    quize_points INTEGER DEFAULT 0,
                    quize_1 INTEGER DEFAULT 0,
                    quize_2 INTEGER DEFAULT 0,
                    quize_3 INTEGER DEFAULT 0,
                    quize_4 INTEGER DEFAULT 0,
                    quize_5 INTEGER DEFAULT 0
                )
            """)
    except Exception as err:
        print(f"Error creating users table: {err}")


async def add_user(user_id: int, username: str):
    """
    Add a new user to the database.

    Args:
        user_id (int): Telegram user ID.
        username (str): Telegram username (without '@').

    Returns:
        bool: True if a new user was inserted, False if the user already existed.

    Raises:
        Exception: Re-raises any unexpected database errors.
    """
    try:
        async with db_manager.get_connection() as conn:
            cursor = await conn.execute(
                "SELECT 1 FROM users WHERE username = ?",
                (username,)
            )
            if await cursor.fetchone():
                print(f"⚠️ {username} уже в базе.")
                return False

            await conn.execute(
                "INSERT INTO users (username, id) VALUES (?, ?)",
                (username, user_id)
            )
            await conn.commit()
            logging.info(
                "Пользователь %s (id: %s) добавлен/обновлен", username, user_id)
            return True

    except Exception as err:
        logging.error("Ошибка добавления пользователя: %s", err, exc_info=True)
        raise


async def get_user_by_username(username: str):
    """
    Retrieve a user's record by their username.

    Args:
        username (str): Telegram username (without '@').

    Returns:
        dict: A mapping of column names to values if the user exists.
        None: If no such user is found or on error.
    """
    try:
        async with db_manager.get_connection() as conn:
            cursor = await conn.execute(
                "SELECT * FROM users WHERE username = ?",
                (username,)
            )
            row = await cursor.fetchone()
            return dict(row) if row else None
    except Exception as err:
        logging.error("Ошибка получения пользователя: %s", err, exc_info=True)
        return None


async def save_users_to_excel():
    """
    Export all users to an Excel file named "users.xlsx".

    The first row will contain headers and be styled in bold white-on-blue.
    Subsequent rows will contain each user’s data in the same order
    as returned by get_all_users().

    Returns:
        str: The filename ("users.xlsx") on success.
        None: If there are no users or on error.
    """
    try:
        users = await get_all_users()
        if not users:
            return None

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = USERS_TITLE

        sheet.append(USERS_HEADERS)
        for user in users:
            sheet.append(list(user.values()))

        header_fill = PatternFill(start_color=USERS_HEADER_COLOR,
                                  end_color=USERS_HEADER_COLOR, fill_type=USERS_FILL_TYPE)
        header_font = Font(bold=True, color=USERS_HEADER_FONT_COLOR)

        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font

        workbook.save(USERS_FILENAME)
        return USERS_FILENAME
    except Exception as err:
        print(f"Error saving to Excel: {err}")
        return None


async def get_all_users():
    """
    Fetch all users from the database.

    Returns:
        list[dict]: A list of user records (empty list on error).
    """
    try:
        async with db_manager.get_connection() as conn:
            cursor = await conn.execute("SELECT * FROM users")
            rows = await cursor.fetchall()
            return [dict(row) for row in rows]
    except Exception as err:
        logging.error("Error getting all users: %s", err, exc_info=True)
        return []


async def start_quest(username: str):
    """
    Mark a user’s quest as started.

    Args:
        username (str): Telegram username.
    """
    try:
        async with db_manager.get_connection() as conn:
            await conn.execute(
                "UPDATE users SET quest_started = 1 WHERE username = ?",
                (username, )
            )
    except Exception as err:
        print(f"Error starting quest: {err}")


async def is_quest_started(username: str):
    """
    Check whether a user has started the quest.

    Args:
        username (str): Telegram username.

    Returns:
        bool: True if quest_started == 1, False otherwise or on error.
    """
    try:
        async with db_manager.get_connection() as conn:
            async with conn.execute(
                "SELECT quest_started FROM users WHERE username = ?",
                    (username, )) as cursor:
                result = await cursor.fetchone()
                return result[0] == 1
    except Exception as err:
        print(f"Error checking quest start: {err}")
        return False


async def check_points(username: str):
    """
    Retrieve the total quest points for a user.

    Args:
        username (str): Telegram username.

    Returns:
        int: quest_points or 0 if none / on error.
    """
    try:
        async with db_manager.get_connection() as conn:
            async with conn.execute(
                "SELECT quest_points FROM users WHERE username = ?",
                (username,)
            ) as cursor:
                result = await cursor.fetchone()
                return result[0] if result else 0
    except Exception as err:
        print(f"Error checking points: {err}")
        return 0


async def update_merch_points(username: str, points: int):
    """
    Deduct the given amount of points from a user when they redeem merch.

    Args:
        username (str): Telegram username.
        points (int): Number of points to subtract.
    """
    try:
        async with db_manager.get_connection() as conn:
            await conn.execute(
                "UPDATE users SET quest_points = quest_points - ? "
                "WHERE username = ?",
                (points, username)
            )
    except Exception as err:
        print(f"Error updating merch points: {err}")


async def check_st_points(username: str, station: int):
    """
    Get the points a user has earned at a specific quest station.

    Args:
        username (str): Telegram username.
        station (int): Station number (1–11).

    Returns:
        int | None: The points at that station, or None on error.
    """
    try:
        async with db_manager.get_connection() as conn:
            async with conn.execute(
                f"SELECT quest{station}_points FROM users WHERE username = ?",
                (username,)
            ) as cursor:
                result = await cursor.fetchone()
                return result[0] if result else None
    except Exception as err:
        print(f"Error checking station points: {err}")
        return 0


async def check_quiz_points(username: str, num: int):
    """
    Get the points a user has earned on a specific quiz question.

    Args:
        username (str): Telegram username.
        num (int): Quiz question number (1–5).

    Returns:
        int | None: The points on that quiz question, or None on error.
    """
    try:
        async with db_manager.get_connection() as conn:
            async with conn.execute(f"SELECT quize_{num} FROM users WHERE username = ?",
                                    (username, )
                                    ) as cursor:
                result = await cursor.fetchone()
                return result[0] if result else None
    except Exception as err:
        print(f"Error checking quiz points: {err}")
        return 0


async def update_user_queststation(username: str):
    """
    Recompute and store a user's current quest station based on
    how many individual station‐point columns are > 0.

    Args:
        username (str): Telegram username.
    """
    try:
        async with db_manager.get_connection() as conn:
            await conn.execute("""
                UPDATE users
                SET quest_station = (
                    (quest1_points > 0) +
                    (quest2_points > 0) +
                    (quest3_points > 0) +
                    (quest4_points > 0) +
                    (quest5_points > 0) +
                    (quest6_points > 0) +
                    (quest7_points > 0) +
                    (quest8_points > 0) +
                    (quest9_points > 0) +
                    (quest10_points > 0) +
                    (quest11_points > 0)
                )
                WHERE username = ?
                """, (username,))
    except Exception as err:
        print(f"Error updating quest station: {err}")


async def is_quest_finished(username: str):
    """
    Check whether a user has finished the quest by determining
    if they have redeemed any merch.

    Args:
        username (str): Telegram username.

    Returns:
        bool: True if the user has any merch, False otherwise.
    """
    return await is_got_any_merch(username)


async def is_quiz_finished(username: str):
    """
    Check whether a user has completed all quiz questions.

    Args:
        username (str): Telegram username.

    Returns:
        bool: True if all quize_1‥quize_5 > 0, False otherwise or on error.
    """
    try:
        async with db_manager.get_connection() as conn:
            result = await conn.fetchrow("""
                SELECT quize_1, quize_2, quize_3, quize_4, quize_5
                FROM users WHERE username = ?
            """, username)
            return all(score > 0 for score in result.values()) if result else False
    except Exception as err:
        print(f"Error checking quiz completion: {err}")
        return False


async def count_active_quests():
    """
    Count how many users have started the quest.

    Returns:
        int: Number of users where quest_started == 1, or 0 on error.
    """
    try:
        async with db_manager.get_connection() as conn:
            async with conn.execute(
                "SELECT COUNT(*) FROM users WHERE quest_started = 1"
            ) as cursor:
                result = await cursor.fetchone()
                return result[0] if result else 0
    except Exception as err:
        print(f"Error counting active quests: {err}")
        return 0


async def count_finished_quests():
    """
    Count how many users have finished the quest by redeeming merch.

    Returns:
        int: Sum of is_got_any_merch(username) over all users, or 0 on error.
    """
    try:
        users = await get_all_users()
        return sum([await is_got_any_merch(user['username']) for user in users])
    except Exception as err:
        print(f"Error counting finished quests: {err}")
        return 0


async def update_user_points(username: str, admin_num: int, points: int):
    """
    Add quest points to both the specific station and the total.

    Args:
        username (str): Telegram username.
        admin_num (int): Station number to update.
        points (int): Points to add.
    """
    try:
        async with db_manager.get_connection() as conn:
            await conn.execute(
                f"UPDATE users SET quest{admin_num}_points = quest{admin_num}_points + ? "
                "WHERE username = ?",
                (points, username)
            )
            await conn.execute(
                "UPDATE users SET quest_points = quest_points + ? "
                "WHERE username = ?",
                (points, username)
            )
    except Exception as err:
        print(f"Error updating user points: {err}")


async def update_quize_points(username: str, num: int):
    """
    Add one point to a specific quiz question and to the total quiz points.

    Args:
        username (str): Telegram username.
        num (int): Quiz question number (1–5).
    """
    try:
        async with db_manager.get_connection() as conn:
            await conn.execute(
                f"UPDATE users SET quize_{num} = quize_{num} + 1 "
                "WHERE username = ?",
                (username, )
            )
            await conn.execute(
                "UPDATE users SET quize_points = quize_points + 1 "
                "WHERE username = ?",
                (username, )
            )
    except Exception as err:
        print(f"Error updating quiz points: {err}")


async def is_finished_quiz(username: str, num: int):
    """
    Check whether a particular quiz question has been answered at least once.

    Args:
        username (str): Telegram username.
        num (int): Quiz question number.

    Returns:
        bool: True if quize_<num> > 0, False otherwise or on error.
    """
    try:
        async with db_manager.get_connection() as conn:
            async with conn.execute(
                f"SELECT quize_{num} FROM users WHERE username = ?",
                (username,)
            ) as cursor:
                result = await cursor.fetchone()
                return result[0] > 0 if result else False
    except Exception as err:
        print(f"Database error checking quiz finish: {err}")
        return False
