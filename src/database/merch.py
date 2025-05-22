import datetime

import openpyxl
from openpyxl.styles import Border, Font, NamedStyle, PatternFill, Side

from constants import (MERCH_BORDER_STYLE, MERCH_FILENAME, MERCH_HEADER_COLOR,
                       MERCH_HEADER_FONT_COLOR, MERCH_HEADER_NAME, MERCH_TYPES)
from src.database.base import db_manager


async def create_merch_table():
    """
    Create the `merch` table if it does not already exist.

    The table schema is:
        - username TEXT UNIQUE
        - one INTEGER column per merch type, defaulting to 0.

    This function should be called once at startup.
    """
    columns_sql = ", ".join(
        f'"{name}" INTEGER DEFAULT 0'
        for name in MERCH_TYPES
    )
    async with db_manager.get_connection() as conn:
        await conn.execute(f"""
            CREATE TABLE IF NOT EXISTS merch (
                username TEXT UNIQUE,
                {columns_sql}
            )
        """)


async def is_valid_column(column_name):
    """
    Check whether `column_name` exists in the `merch` table.

    Args:
        column_name (str): Name of the column to validate.

    Returns:
        bool: True if the column exists (case-insensitive), False otherwise.
    """
    async with db_manager.get_connection() as conn:
        cursor = await conn.execute("PRAGMA table_info(merch)")
        columns = await cursor.fetchall()
        return any(row[1].strip().lower() == column_name.strip().lower() for row in columns)


async def got_merch(username, merch_type):
    """
    Check if a user has already received a specific merch item.

    If the user row does not exist, it will be created with all flags = 0.

    Args:
        username (str): Telegram username.
        merch_type (str): Column name of the merch item.

    Returns:
        bool: True if the user’s column value == 1, False otherwise.

    Raises:
        ValueError: If `merch_type` is not a valid column.
    """
    if not await is_valid_column(merch_type):
        raise ValueError(f"Недопустимое имя колонки: {merch_type}")

    async with db_manager.get_connection() as conn:
        await conn.execute(
            "INSERT OR IGNORE INTO merch (username) VALUES (?)",
            (username,)
        )
        await conn.commit()
        cursor = await conn.execute(
            f'SELECT "{merch_type}" FROM merch WHERE username = ?',
            (username,)
        )
        result = await cursor.fetchone()

        if result:
            return result[merch_type] == 1
        return False


async def give_merch(username, merch_type):
    """
    Mark that a user has received a specific merch item.

    Creates the user row if necessary, then sets the given column to 1.

    Args:
        username (str): Telegram username.
        merch_type (str): Column name of the merch item.

    Raises:
        ValueError: If `merch_type` is not a valid column.
    """
    if not await is_valid_column(merch_type):
        raise ValueError(f"Недопустимое имя колонки: {merch_type}")

    async with db_manager.get_connection() as conn:
        await conn.execute(
            "INSERT INTO merch (username) VALUES (?) ON CONFLICT (username) DO NOTHING",
            (username, )
        )
        await conn.execute(
            f'UPDATE merch SET "{merch_type}" = 1 WHERE username = ?',
            (username,)
        )
        await conn.commit()


async def is_got_merch(username):
    """
    Determine if a user has received *all* defined merch items.

    Args:
        username (str): Telegram username.

    Returns:
        bool: True if every merch-type column for this user == 1, False otherwise.
    """
    async with db_manager.get_connection() as conn:
        await conn.execute(
            "INSERT INTO merch (username) VALUES (?) ON CONFLICT(username) DO NOTHING",
            (username,)
        )
        await conn.commit()

        cursor = await conn.execute("PRAGMA table_info(merch)")
        rows = await cursor.fetchall()
        numeric_columns = [row[1] for row in rows if row[1] != 'username']

        if not numeric_columns:
            return False

        check_conditions = " AND ".join(
            [f'"{col}" = 1' for col in numeric_columns]
        )
        query = f"""
            SELECT CASE WHEN ({check_conditions}) THEN 1 ELSE 0 END
            FROM merch
            WHERE username = ?
        """
        cursor2 = await conn.execute(query, (username,))
        result = await cursor2.fetchone()
        return bool(result[0]) if result else False


async def is_got_any_merch(username):
    """
    Determine if a user has received *any* merch items.

    Args:
        username (str): Telegram username.

    Returns:
        bool: True if the sum of all merch-type columns > 0, False otherwise.
    """
    async with db_manager.get_connection() as conn:
        await conn.execute(
            "INSERT INTO merch (username) VALUES (?) ON CONFLICT(username) DO NOTHING",
            (username,)
        )

        cursor = await conn.execute("PRAGMA table_info(merch)")
        columns = await cursor.fetchall()
        numeric_columns = [col[1] for col in columns if col[1] != 'username']

        if not numeric_columns:
            return False

        sum_query = " + ".join([f'"{col}"' for col in numeric_columns])
        cursor2 = await conn.execute(
            f"SELECT ({sum_query}) > 0 FROM merch WHERE username = ?",
            (username,)
        )
        result = await cursor2.fetchone()
        return bool(result[0]) if result else False


async def add_column(column_name, column_type="INTEGER DEFAULT 0"):
    """
    Dynamically add a new merch-type column to the `merch` table.

    If the column already exists, does nothing.

    Args:
        column_name (str): Name of the new column.
        column_type (str): SQL type and default (e.g. "INTEGER DEFAULT 0").
    """
    async with db_manager.get_connection() as conn:
        cursor = await conn.execute("PRAGMA table_info(merch)")
        columns = await cursor.fetchall()
        column_names = [col[1] for col in columns]

        if column_name not in column_names:
            await conn.execute(
                f'ALTER TABLE merch ADD COLUMN "{column_name}" {column_type}'
            )
            print(f"Столбец '{column_name}' добавлен в таблицу.")
        else:
            print(f"Столбец '{column_name}' уже существует.")
        await conn.commit()


async def get_all_merch():
    """
    Retrieve all rows from the `merch` table.

    Returns:
        List[dict]: A list of dicts mapping column names to values.
    """
    async with db_manager.get_connection() as conn:
        cursor = await conn.execute("SELECT * FROM merch")
        rows = await cursor.fetchall()
        return [dict(row) for row in rows]


async def get_table_columns(table_name):
    """
    List the column names of any database table.

    Args:
        table_name (str): Name of the table.

    Returns:
        List[str]: Column names in definition order.
    """
    async with db_manager.get_connection() as conn:
        cursor = await conn.execute(f"PRAGMA table_info({table_name})")
        columns = await cursor.fetchall()
        return [col[1] for col in columns]


async def save_merch_to_excel():
    """
    Export the entire merch table to a formatted Excel file.

    - Adds a header row with bold white text on blue fill.
    - Converts datetime values to strings.
    - Auto-sizes columns to fit content.
    - Saves to "merch.xlsx" in the current directory.

    Returns:
        str: Filename ("merch.xlsx") if merch rows exist, otherwise None.
    """

    columns = await get_table_columns('merch')
    if not await get_all_merch():
        return None

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Мерч"
    header_style = NamedStyle(name=MERCH_HEADER_NAME)
    header_style.font = Font(bold=True, color=MERCH_HEADER_FONT_COLOR)
    header_style.fill = PatternFill("solid", fgColor=MERCH_HEADER_COLOR)
    header_style.border = Border(bottom=Side(style=MERCH_BORDER_STYLE))
    for col_num, column_title in enumerate(columns, 1):
        cell = sheet.cell(row=1, column=col_num, value=column_title)
        cell.style = header_style
    for row_num, item in enumerate(await get_all_merch(), 2):
        for col_num, column_name in enumerate(columns, 1):
            value = item.get(column_name)
            if isinstance(value, datetime.datetime):
                value = value.strftime("%Y-%m-%d %H:%M:%S")
            elif value is None:
                value = ""
            sheet.cell(row=row_num, column=col_num, value=value)
    for column in sheet.columns:
        max_length = 0
        for cell in column:
            try:
                length = len(str(cell.value))
                if length > max_length:
                    max_length = length
            except Exception:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column[0].column_letter].width = adjusted_width
    workbook.save(MERCH_FILENAME)
    workbook.close()
    return MERCH_FILENAME
