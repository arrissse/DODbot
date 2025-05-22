import os

import aiosqlite
import pytest
import pytest_asyncio
from openpyxl import load_workbook

from database.base import db_manager
from src.database import merch as merch_module


@pytest_asyncio.fixture(autouse=True)
async def in_memory_db(monkeypatch):
    """
    Создаёт временную in-memory базу данных SQLite и заменяет
    `get_connection` у `db_manager`, чтобы использовать её.
    После выполнения тестов удаляет файл `merch.xlsx`, если он был создан.

    Args:
        monkeypatch: Fixture for safely patching attributes.

    Yields:
        None: Control is passed to the test function.
    """
    conn = await aiosqlite.connect(":memory:")
    conn.row_factory = aiosqlite.Row

    class InMemCtx:
        """Контекстный менеджер для подмены подключения к SQLite."""
        async def __aenter__(self):
            return conn

        async def __aexit__(self, exc_type, exc, exc_tb):
            pass

    monkeypatch.setattr(db_manager, "get_connection", InMemCtx())

    await merch_module.create_merch_table()

    yield

    await conn.close()

    try:
        os.remove("merch.xlsx")
    except FileNotFoundError:
        pass


@pytest.mark.asyncio
async def test_create_and_validate_columns():
    """Проверяет, что в таблице есть нужные колонки и валидация колонок работает корректно."""
    cols = await merch_module.get_table_columns("merch")
    assert "username" in cols
    for name in ["Раскрасить шоппер", "Шоппер", "Раскрасить футболку", "Футболка", "Пауэрбанк"]:
        assert name in cols

    assert await merch_module.is_valid_column("Шоппер") is True
    assert await merch_module.is_valid_column("nonexistent") is False


@pytest.mark.asyncio
async def test_got_and_give_merch_flags():
    """Проверяет логику получения и выдачи мерча конкретному пользователю."""
    user = "alice"

    assert await merch_module.got_merch(user, "Шоппер") is False

    await merch_module.give_merch(user, "Шоппер")
    assert await merch_module.got_merch(user, "Шоппер") is True

    assert await merch_module.got_merch(user, "Футболка") is False

    with pytest.raises(ValueError):
        await merch_module.got_merch(user, "invalid")


@pytest.mark.asyncio
async def test_is_got_merch_all_vs_any():
    """Проверяет is_got_any_merch и is_got_merch (все vs хотя бы один)."""
    user = "bob"
    assert await merch_module.is_got_any_merch(user) is False
    assert await merch_module.is_got_merch(user) is False

    await merch_module.give_merch(user, "Раскрасить шоппер")
    assert await merch_module.is_got_any_merch(user) is True
    assert await merch_module.is_got_merch(user) is False

    for col in await merch_module.get_table_columns("merch"):
        if col != "username":
            await merch_module.give_merch(user, col)
    assert await merch_module.is_got_any_merch(user) is True
    assert await merch_module.is_got_merch(user) is True


@pytest.mark.asyncio
async def test_add_column_and_reflect_in_schema():
    """Проверяет добавление новой колонки и её наличие в схеме."""
    await merch_module.add_column("НовыйМерч")
    cols = await merch_module.get_table_columns("merch")
    assert "НовыйМерч" in cols

    user = "charlie"
    await merch_module.give_merch(user, "НовыйМерч")
    assert await merch_module.got_merch(user, "НовыйМерч") is True


@pytest.mark.asyncio
async def test_get_all_merch_and_to_dicts():
    """Проверяет, что get_all_merch возвращает корректные данные в виде списка словарей."""
    await merch_module.give_merch("u1", "Шоппер")
    await merch_module.give_merch("u2", "Футболка")

    all_data = await merch_module.get_all_merch()
    assert isinstance(all_data, list)
    assert any(item["username"] == "u1" and item["Шоппер"]
               == 1 for item in all_data)
    assert any(item["username"] == "u2" and item["Футболка"]
               == 1 for item in all_data)


@pytest.mark.asyncio
async def test_save_merch_to_excel_and_content():
    """Проверяет экспорт мерча в Excel и корректность содержимого."""
    await merch_module.give_merch("dave", "Пауэрбанк")
    filename = await merch_module.save_merch_to_excel()
    assert filename == "merch.xlsx"
    assert os.path.exists(filename)

    workbook = load_workbook(filename)
    worksheet = workbook.active
    headers = [c.value for c in worksheet[1]]
    expected = await merch_module.get_table_columns("merch")
    assert headers == expected

    data_rows = list(worksheet.iter_rows(min_row=2, values_only=True))
    assert any(row[0] == "dave" for row in data_rows)
