from unittest.mock import AsyncMock

import aiosqlite
import pytest
import pytest_asyncio
from openpyxl import load_workbook

from database.base import db_manager
from src.database import admin as admin_module

pytest_plugins = ("pytest_asyncio",)


class InMemoryConn:
    def __init__(self, conn):
        self._conn = conn

    async def __aenter__(self):
        return self._conn

    async def __aexit__(self, exc_type, exc, tb):
        pass


@pytest_asyncio.fixture(autouse=True)
async def in_memory_db(monkeypatch):
    conn = await aiosqlite.connect(":memory:")
    conn.row_factory = aiosqlite.Row

    def get_connection():
        return InMemoryConn(conn)

    monkeypatch.setattr(db_manager, "get_connection", get_connection)
    await admin_module.create_admins_table()
    yield
    await conn.close()

@pytest.mark.asyncio
async def test_update_admin_questnum_and_get_by_username():
    await admin_module.add_admin('@station_admin', 2)
    record = await admin_module.get_admin_by_username('@station_admin')
    assert record[2] == 0
    await admin_module.update_admin_questnum('@station_admin', 5)
    record2 = await admin_module.get_admin_by_username('@station_admin')
    assert record2[2] == 5


@pytest.mark.asyncio
async def test_get_admin_level_not_found():
    level = await admin_module.get_admin_level('@missing')
    assert level == 0


@pytest.mark.asyncio
async def test_save_admins_to_excel(tmp_path):
    await admin_module.add_admin('@u1', 0)
    await admin_module.add_admin('@u2', 1)
    await admin_module.update_admin_questnum('@u2', 3)

    file = await admin_module.save_admins_to_excel()
    assert isinstance(file, str)
    wb = load_workbook(file)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    assert headers == ['Username', 'Level', 'Station']
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert ('@u1', 0, 0) in rows
    assert ('@u2', 1, 3) in rows


@pytest.mark.asyncio
async def test_init_admins_calls_add_and_update(monkeypatch):
    add_mock = AsyncMock()
    upd_mock = AsyncMock()
    get_admin_mock = AsyncMock(
        side_effect=lambda username: ('@' + username, 0, 0))

    monkeypatch.setattr(admin_module, 'add_admin', add_mock)
    monkeypatch.setattr(admin_module, 'update_admin_questnum', upd_mock)
    monkeypatch.setattr(admin_module, 'get_admin_by_username', get_admin_mock)

    await admin_module.init_admins()

    add_mock.assert_any_await('@arrisse', 0)
    add_mock.assert_any_await('@Nikita_Savel', 0)
    add_mock.assert_any_await('@thelasttime111th', 2)
    upd_mock.assert_any_await('@thelasttime111th', 9)
    add_mock.assert_any_await('@kson27', 1)
