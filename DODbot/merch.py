import openpyxl
from database import db_manager
import datetime
import logging

async def create_merch_table():
    async with db_manager.get_connection() as conn:
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS merch (
                username TEXT UNIQUE,
                "Раскрасить шоппер" INTEGER DEFAULT 0,
                "Шоппер" INTEGER DEFAULT 0,
                "Раскрасить футболку" INTEGER DEFAULT 0,
                "Футболка" INTEGER DEFAULT 0,
                "Пауэрбанк" INTEGER DEFAULT 0
            )
        """)


async def is_valid_column(column_name: str) -> bool:
    async with db_manager.get_connection() as conn:
        cursor = await conn.execute("PRAGMA table_info(merch)")
        columns = await cursor.fetchall()
        # row[1] содержит имя столбца
        return any(row[1].strip().lower() == column_name.strip().lower() for row in columns)


async def got_merch(username: str, merch_type: str) -> bool:
    if not await is_valid_column(merch_type):
        raise ValueError(f"Недопустимое имя колонки: {merch_type}")

    async with db_manager.get_connection() as conn:
        await conn.execute(
            "INSERT OR IGNORE INTO merch (username) VALUES (?)",
            (username,)
        )
        await conn.commit()
        cursor = await conn.execute(
            f'SELECT "{merch_type}" FROM merch WHERE username = ?',
            (username,)
        )
        result = await cursor.fetchone()

        if result:
            return result[merch_type] == 1
        return False


async def give_merch(username: str, merch_type: str):
    if not await is_valid_column(merch_type):
        raise ValueError(f"Недопустимое имя колонки: {merch_type}")

    async with db_manager.get_connection() as conn:
        await conn.execute(
            "INSERT INTO merch (username) VALUES (?) ON CONFLICT (username) DO NOTHING",
            (username, )
        )
        await conn.execute(
            f'UPDATE merch SET "{merch_type}" = 1 WHERE username = ?',
            (username,)
        )
        await conn.commit()


async def is_got_merch(username: str) -> bool:
    async with db_manager.get_connection() as conn:
        await conn.execute(
            "INSERT INTO merch (username) VALUES (?) ON CONFLICT(username) DO NOTHING",
            (username,)
        )
        await conn.commit()

        cursor = await conn.execute("PRAGMA table_info(merch)")
        rows = await cursor.fetchall()
        numeric_columns = [row[1] for row in rows if row[1] != 'username']

        if not numeric_columns:
            return False

        check_conditions = " AND ".join(
            [f'"{col}" = 1' for col in numeric_columns]
        )
        query = f"""
            SELECT CASE WHEN ({check_conditions}) THEN 1 ELSE 0 END
            FROM merch 
            WHERE username = ?
        """
        cursor2 = await conn.execute(query, (username,))
        result = await cursor2.fetchone()
        return bool(result[0]) if result else False


async def is_got_any_merch(username: str) -> bool:
    async with db_manager.get_connection() as conn:
        await conn.execute(
            "INSERT INTO merch (username) VALUES (?) ON CONFLICT(username) DO NOTHING",
            (username,)
        )

        cursor = await conn.execute("PRAGMA table_info(merch)")
        columns = await cursor.fetchall()
        numeric_columns = [col[1] for col in columns if col[1] != 'username']

        if not numeric_columns:
            return False

        sum_query = " + ".join([f'"{col}"' for col in numeric_columns])
        cursor2 = await conn.execute(
            f"SELECT ({sum_query}) > 0 FROM merch WHERE username = ?",
            (username,)
        )
        result = await cursor2.fetchone()
        return bool(result[0]) if result else False


async def add_column(column_name: str, column_type: str = "INTEGER DEFAULT 0"):
    async with db_manager.get_connection() as conn:
        cursor = await conn.execute("PRAGMA table_info(merch)")
        columns = await cursor.fetchall()
        column_names = [col[1] for col in columns]

        if column_name not in column_names:
            await conn.execute(
                f'ALTER TABLE merch ADD COLUMN "{column_name}" {column_type}'
            )
            print(f"Столбец '{column_name}' добавлен в таблицу.")
        else:
            print(f"Столбец '{column_name}' уже существует.")
        await conn.commit()


async def get_all_merch():
    async with db_manager.get_connection() as conn:
        cursor = await conn.execute("SELECT * FROM merch")
        rows = await cursor.fetchall()
        return [dict(row) for row in rows]


async def get_table_columns(table_name: str):
    async with db_manager.get_connection() as conn:
        cursor = await conn.execute(f"PRAGMA table_info({table_name})")
        columns = await cursor.fetchall()
        return [col[1] for col in columns]


async def save_merch_to_excel():
    merch = await get_all_merch()
    columns = await get_table_columns('merch')
    if not merch:
        return None
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, NamedStyle
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Мерч"
    header_style = NamedStyle(name="header")
    header_style.font = Font(bold=True, color="FFFFFF")
    header_style.fill = PatternFill("solid", fgColor="4F81BD")
    header_style.border = Border(bottom=Side(style="thin"))
    for col_num, column_title in enumerate(columns, 1):
        cell = sheet.cell(row=1, column=col_num, value=column_title)
        cell.style = header_style
    for row_num, item in enumerate(merch, 2):
        for col_num, column_name in enumerate(columns, 1):
            value = item.get(column_name)
            if isinstance(value, datetime):
                value = value.strftime("%Y-%m-%d %H:%M:%S")
            elif value is None:
                value = ""
            sheet.cell(row=row_num, column=col_num, value=value)
    # Автоматическая настройка ширины столбцов
    for column in sheet.columns:
        max_length = 0
        for cell in column:
            try:
                length = len(str(cell.value))
                if length > max_length:
                    max_length = length
            except Exception:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column[0].column_letter].width = adjusted_width
    filename = "merch.xlsx"
    workbook.save(filename)
    workbook.close()
    return filename
