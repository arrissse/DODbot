import openpyxl
from openpyxl.styles import Font, PatternFill
from merch import is_got_any_merch
from database import db_manager
import asyncpg
import logging


async def create_users_table():
    """Создание таблицы пользователей (асинхронная версия)"""
    try:
        async with db_manager.get_connection() as conn:
            await conn.execute("""
                CREATE TABLE IF NOT EXISTS users (
                    id INTEGER PRIMARY KEY,
                    username TEXT,
                    quest_started INTEGER DEFAULT 0,
                    quest1_points INTEGER DEFAULT 0,
                    quest2_points INTEGER DEFAULT 0,
                    quest3_points INTEGER DEFAULT 0,
                    quest4_points INTEGER DEFAULT 0,
                    quest5_points INTEGER DEFAULT 0,
                    quest6_points INTEGER DEFAULT 0,
                    quest7_points INTEGER DEFAULT 0,
                    quest8_points INTEGER DEFAULT 0,
                    quest9_points INTEGER DEFAULT 0,
                    quest10_points INTEGER DEFAULT 0,
                    quest11_points INTEGER DEFAULT 0,
                    quest_station INTEGER DEFAULT 0,
                    quest_points INTEGER DEFAULT 0,
                    quize_points INTEGER DEFAULT 0,
                    quize_1 INTEGER DEFAULT 0,
                    quize_2 INTEGER DEFAULT 0,
                    quize_3 INTEGER DEFAULT 0,
                    quize_4 INTEGER DEFAULT 0,
                    quize_5 INTEGER DEFAULT 0
                )
            """)
    except Exception as e:
        print(f"Error creating users table: {e}")


async def add_user(user_id: int, username: str):
    """Добавление пользователя (асинхронная версия)"""
    try:
        async with db_manager.get_connection() as conn:
            await conn.execute(
                "INSERT INTO users (id, username) VALUES (?, ?) "
                "ON CONFLICT (id) DO NOTHING",
                (user_id, username)
            )
            await conn.commit()

            logging.info(
                f"Пользователь {username} (id: {user_id}) добавлен/обновлен")

    except Exception as e:
        # Подробное логирование ошибки с трейсбэком
        logging.error(f"Ошибка добавления пользователя: {e}", exc_info=True)
        raise


async def get_user_by_username(username: str) -> dict:
    """Получение пользователя по username"""
    try:
        async with db_manager.get_connection() as conn:
            cursor = await conn.execute(
                "SELECT * FROM users WHERE username = ?",
                (username,)
            )
            row = await cursor.fetchone()
            return dict(row) if row else None
    except Exception as e:
        logging.error(f"Ошибка получения пользователя: {e}", exc_info=True)
        return None


async def save_users_to_excel() -> str:
    """Экспорт в Excel (асинхронная версия)"""
    try:
        users = await get_all_users()
        if not users:
            return None

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Пользователи"

        headers = [
            "ID", "Username", "Участие в квесте", "РТ СТАНЦИЯ", "ЛФИ СТАНЦИЯ",
            "ФАКТ СТАНЦИЯ", "ФЭФМ СТАНЦИЯ", "ФПМИ СТАНЦИЯ", "ФБМФ СТАНЦИЯ",
            "КНТ СТАНЦИЯ", "ФБВТ СТАНЦИЯ", "ВШПИ СТАНЦИЯ", "ВШМ СТАНЦИЯ",
            "ПИШ РПИ СТАНЦИЯ", "Количество пройденных станций",
            "Квест сумма баллов", "Квиз", "1", "2", "3", "4", "5"
        ]

        sheet.append(headers)
        for user in users:
            sheet.append(list(user.values()))

        # Стилизация
        header_fill = PatternFill(start_color="4F81BD",
                                  end_color="4F81BD", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font

        filename = "users.xlsx"
        workbook.save(filename)
        return filename
    except Exception as e:
        print(f"Error saving to Excel: {e}")
        return None


async def get_all_users() -> list:
    try:
        async with db_manager.get_connection() as conn:
            cursor = await conn.execute("SELECT * FROM users")
            logging.info(f"get_all_users: {await cursor.fetchall()}")
            return await cursor.fetchall()
    except Exception as e:
        logging.info(f"Error getting all users: {e}")
        return []


async def start_quest(username: str):
    """Старт квеста (асинхронная версия)"""
    try:
        async with db_manager.get_connection() as conn:
            await conn.execute(
                "UPDATE users SET quest_started = 1 WHERE username = ?",
                (username, )
            )
    except Exception as e:
        print(f"Error starting quest: {e}")


async def finish_quest(username: str):
    """Завершение квеста (асинхронная версия)"""
    try:
        async with db_manager.get_connection() as conn:
            await conn.execute(
                "UPDATE users SET quest_station = 11 WHERE username = ?",
                (username, )
            )
    except Exception as e:
        print(f"Error finishing quest: {e}")


async def is_quest_started(username: str) -> bool:
    """Проверка старта квеста (асинхронная версия)"""
    try:
        async with db_manager.get_connection() as conn:
            async with conn.execute(
                "SELECT quest_started FROM users WHERE username = ?",
                    (username, )) as cursor:
                result = await cursor.fetchone()
                return result[0] == 1
    except Exception as e:
        print(f"Error checking quest start: {e}")
        return False


async def check_points(username: str) -> int:
    """Проверка баллов (асинхронная версия)"""
    try:
        async with db_manager.get_connection() as conn:
            async with conn.execute((
                "SELECT quest_points FROM users WHERE username = ?",
                (username,)
            )) as cursor:
                result = await cursor.fetchone()
                return result[0] or 0
    except Exception as e:
        print(f"Error checking points: {e}")
        return 0


async def update_merch_points(username: str, points: int):
    """Обновление баллов мерча (асинхронная версия)"""
    try:
        async with db_manager.get_connection() as conn:
            await conn.execute(
                "UPDATE users SET quest_points = quest_points - ? "
                "WHERE username = ?",
                points, username
            )
    except Exception as e:
        print(f"Error updating merch points: {e}")


async def check_st_points(username: str, station: int) -> int:
    """Проверка баллов станции (асинхронная версия)"""
    try:
        async with db_manager.get_connection() as conn:
            async with conn.execute(
                f"SELECT quest{station}_points FROM users WHERE username = ?",
                (username,)
            ) as cursor:
                result = await cursor.fetchone()
                return result[0] if result else None
    except Exception as e:
        print(f"Error checking station points: {e}")
        return 0


async def check_quiz_points(username: str, num: int) -> int:
    """Проверка баллов квиза (асинхронная версия)"""
    try:
        async with db_manager.get_connection() as conn:
            async with conn.execute(f"SELECT quize_{num} FROM users WHERE username = ?",
                                    (username, )
            ) as cursor:
                result=await cursor.fetchone()
                return result[0] if result else None
    except Exception as e:
        print(f"Error checking quiz points: {e}")
        return 0


async def update_user_queststation(username: str):
    """Обновление статуса квеста (асинхронная версия)"""
    try:
        async with db_manager.get_connection() as conn:
            await conn.execute("""
                UPDATE users SET quest_station = (
                    (quest1_points > 0)::integer + 
                    (quest2_points > 0)::integer +
                    (quest3_points > 0)::integer +
                    (quest4_points > 0)::integer +
                    (quest5_points > 0)::integer +
                    (quest6_points > 0)::integer +
                    (quest7_points > 0)::integer +
                    (quest8_points > 0)::integer +
                    (quest9_points > 0)::integer +
                    (quest10_points > 0)::integer +
                    (quest11_points > 0)::integer
                ) WHERE username = ?
            """, username)
    except Exception as e:
        print(f"Error updating quest station: {e}")


async def is_quest_finished(username: str) -> bool:
    """Проверка завершения квеста (асинхронная версия)"""
    return await is_got_any_merch(username)


async def is_quiz_finished(username: str) -> bool:
    """Проверка завершения квизов (асинхронная версия)"""
    try:
        async with db_manager.get_connection() as conn:
            result = await conn.fetchrow("""
                SELECT quize_1, quize_2, quize_3, quize_4, quize_5
                FROM users WHERE username = ?
            """, username)
            return all(score > 0 for score in result.values()) if result else False
    except Exception as e:
        print(f"Error checking quiz completion: {e}")
        return False


async def count_active_quests() -> int:
    """Подсчет активных квестов (асинхронная версия)"""
    try:
        async with db_manager.get_connection() as conn:
            async with conn.execute(
                "SELECT COUNT(*) FROM users WHERE quest_started = 1"
        ) as cursor:
                result = await cursor.fetchone()
                return result[0] if result else 0
    except Exception as e:
        print(f"Error counting active quests: {e}")
        return 0


async def count_finished_quests() -> int:
    """Подсчет завершенных квестов (асинхронная версия)"""
    try:
        users = await get_all_users()
        return sum([await is_got_any_merch(user['username']) for user in users])
    except Exception as e:
        print(f"Error counting finished quests: {e}")
        return 0


async def update_user_points(username: str, admin_num: int, points: int):
    """Обновление баллов пользователя (асинхронная версия)"""
    try:
        async with db_manager.get_connection() as conn:
            await conn.execute(
                f"UPDATE users SET quest{admin_num}_points = quest{admin_num}_points + ? "
                "WHERE username = ?",
                points, username
            )
            await conn.execute(
                "UPDATE users SET quest_points = quest_points + ? "
                "WHERE username = ?",
                points, username
            )
    except Exception as e:
        print(f"Error updating user points: {e}")


async def update_quize_points(username: str, num: int):
    """Обновление баллов квиза (асинхронная версия)"""
    try:
        async with db_manager.get_connection() as conn:
            await conn.execute(
                f"UPDATE users SET quize_{num} = quize_{num} + 1 "
                "WHERE username = ?",
                username
            )
            await conn.execute(
                "UPDATE users SET quize_points = quize_points + 1 "
                "WHERE username = ?",
                username
            )
    except Exception as e:
        print(f"Error updating quiz points: {e}")


async def is_finished_quiz(username: str, num: int) -> bool:
    """Проверка завершения квиза (асинхронная версия)"""
    try:
        async with db_manager.get_connection() as conn:
            async with conn.execute(
                f"SELECT quize_{num} FROM users WHERE username = ?",
                (username,)
            ) as cursor:
                result = await cursor.fetchone()
                return result[0] > 0 if result else False
    except Exception as e:
        print(f"Database error checking quiz finish: {e}")
        return False
    except Exception as e:
        print(f"Unexpected error checking quiz finish: {e}")
        return False
